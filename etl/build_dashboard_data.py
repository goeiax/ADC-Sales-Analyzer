"""Rebuild dashboard-data.json from ALL Neosoft _Raw Monthly Data CSV exports."""
import csv
import json
import os
import re
import traceback
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
# Priority: explicit ADC_RAW_DIR (used by the standalone .exe) → local folder → legacy fallback.
_env_raw = os.environ.get("ADC_RAW_DIR")
if _env_raw:
    RAW_DIR = Path(_env_raw)
else:
    RAW_DIR = SCRIPT_DIR / "_Raw Monthly Data"
    if not RAW_DIR.is_dir():
        RAW_DIR = SCRIPT_DIR.parent / "ADC Files" / "Neosoft Export" / "_Raw Monthly Data"

OUT_DIR = Path(os.environ.get("ADC_DASHBOARD_OUT", str(SCRIPT_DIR)))
OUT_PATH = OUT_DIR / "dashboard-data.json"
HISTORICAL_XLSX = RAW_DIR / "Collection by Invoice 2022, 2023, 2024, 2025. 2026.xlsx"

MONTH_ORDER = [
    ("Jan", "January", "jan"),
    ("Feb", "February", "feb"),
    ("Mar", "March", "mar"),
    ("Apr", "April", "apr"),
    ("May", "May", "may"),
    ("Jun", "June", "jun"),
    ("Jul", "July", "jul"),
    ("Aug", "August", "aug"),
    ("Sep", "September", "sep"),
    ("Oct", "October", "oct"),
    ("Nov", "November", "nov"),
    ("Dec", "December", "dec"),
]

DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

FILE_TYPES = {
    "invoice": "Collection by Invoice {m}.csv",
    "collection": "Item Sold by Collection {m}.csv",
    "collection_list": "Item Sold by Collection List {m}.csv",
    "sales": "Item Sold by Sales {m}.csv",
    "sales_list": "Item Sold by Sales List {m}.csv",
    "commission": "Commission Sales Usage Report {m}.csv",
}


def parse_amount(val):
    try:
        return float(str(val).strip().strip('"').replace(",", "").replace("%", ""))
    except (TypeError, ValueError):
        return 0.0


def parse_pct(val):
    v = parse_amount(val)
    return v if v <= 100 else 0.0


def is_credit(payment_method, item_code):
    pm = str(payment_method).strip().upper()
    ic = str(item_code).strip().upper()
    return "CREDIT VOUCHER" in pm or ic.startswith("CV-")


def is_internal(payment_method):
    return "INTERNAL" in str(payment_method).strip().upper()


def normalize_payment(pm):
    pm = str(pm).strip().upper()
    if not pm:
        return "Other"
    if "BCA" in pm and "EDC" in pm:
        return "EDC BCA"
    if "MANDIRI" in pm:
        return "EDC Mandiri"
    if pm == "CASH" or "CASH" in pm:
        return "Cash"
    if "GOPAY" in pm or "QRIS" in pm:
        return "GoPay/QRIS"
    if "TRF" in pm or "TRANSFER" in pm:
        return "Bank Transfer"
    return pm.title() if pm else "Other"


_MONTH_NUM = {
    "jan": 1, "feb": 2, "peb": 2, "mar": 3, "apr": 4, "may": 5, "mei": 5,
    "jun": 6, "jul": 7, "aug": 8, "agu": 8, "agt": 8, "sep": 9, "oct": 10,
    "okt": 10, "nov": 11, "dec": 12, "des": 12,
}


def parse_invoice_date(raw):
    """Parse a Neosoft invoice date. Tolerant of several export formats:
    '05 Jan 2026' / '5 Mei 2025' / '2026-01-05' / '05/01/2026' / '05.01.2026'."""
    s = str(raw).strip().strip('"')
    if not s:
        return None
    # 1) "DD Mon YYYY" (English or Indonesian month names)
    m = re.search(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", s)
    if m:
        mon = _MONTH_NUM.get(m.group(2)[:3].lower())
        if mon:
            try:
                return datetime(int(m.group(3)), mon, int(m.group(1)))
            except ValueError:
                pass
    # 2) ISO "YYYY-MM-DD"
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    # 3) "DD/MM/YYYY" / "DD-MM-YYYY" / "DD.MM.YYYY" (day-first, Indonesian convention)
    m = re.search(r"(\d{1,2})[/.\-](\d{1,2})[/.\-](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return datetime(y, mo, d)
        except ValueError:
            try:
                return datetime(y, d, mo)  # fall back to month-first if day-first is invalid
            except ValueError:
                pass
    return None


def read_csv(filepath):
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def outlet_label(outlet):
    o = str(outlet).strip()
    if o == "adc SpDVE":
        return "SpDVE"
    if o == "adc SpGK":
        return "SpGK"
    return "Other"


def item_category(code, name="", csv_category=""):
    if csv_category:
        return str(csv_category).strip()
    c = str(code).strip().upper()
    if c.startswith("PKG"):
        return "Package"
    if c.startswith("T-"):
        return "Treatment"
    if c.startswith("P-"):
        return "Product"
    n = str(name).lower()
    if "consult" in n:
        return "Consult"
    return "Other"


def mrn_new_month(mrn, month_key):
    """EM-YYMM — True if MRN registration month matches invoice month key."""
    m = normalize_mrn(mrn)
    if not m.startswith("EM-") or len(m) < 9:
        return False
    yy, mm = m[3:5], m[5:7]
    key_map = {"jan": "01", "feb": "02", "mar": "03", "apr": "04", "may": "05", "jun": "06",
               "jul": "07", "aug": "08", "sep": "09", "oct": "10", "nov": "11", "dec": "12"}
    return key_map.get(month_key) == mm and yy in ("25", "26")


def normalize_mrn(mrn):
    m = str(mrn or "").strip().upper()
    if m in ("", "NONE", "NAN"):
        return ""
    return m


def normalize_name(name):
    n = re.sub(r"\s+", " ", str(name or "").strip().lower())
    return n


def extract_patient_fields(row):
    """Read MRN + name from any Neosoft export row (column name variants)."""
    mrn = ""
    for key in ("Nomor Rekam Medis", "MRN", "Nomor Rekam medis"):
        if key in row and row[key]:
            mrn = normalize_mrn(row[key])
            if mrn:
                break
    name = ""
    for key in ("Patient", "patient"):
        if key in row and row[key]:
            name = str(row[key]).strip()
            if name and name.upper() != "NONE":
                break
    return mrn, name


def extract_phone(row):
    """Pull a phone number from any of the column-name variants Neosoft uses."""
    for key in ("Phone", "Phone Number", "No HP", "No. HP", "No Hp", "Telepon",
                "Mobile", "Handphone", "No. Telepon", "No Telp", "Contact"):
        if key in row and row[key]:
            v = str(row[key]).strip()
            if v and v.upper() not in ("NONE", "NAN", "-"):
                return v
    return ""


class CustomerRegistry:
    """One compiled profile per patient (MRN). Merges invoice, list, sales, commission sources."""

    def __init__(self):
        self._by_mrn = {}
        self._name_mrns = defaultdict(set)

    def resolve(self, mrn, name=""):
        mrn = normalize_mrn(mrn)
        name_clean = str(name or "").strip()
        nk = normalize_name(name_clean) if name_clean else ""

        if mrn:
            if nk:
                self._name_mrns[nk].add(mrn)
            return mrn

        if nk and len(self._name_mrns[nk]) == 1:
            return next(iter(self._name_mrns[nk]))
        return ""

    def touch(self, mrn, name="", source=""):
        mrn = self.resolve(mrn, name)
        if not mrn:
            return None
        name_clean = str(name or "").strip()
        c = self._by_mrn.setdefault(mrn, {
            "mrn": mrn,
            "name": "",
            "names": set(),
            "revenue": 0.0,
            "invoiceVisits": 0,
            "lineItems": 0,
            "commissionUsages": 0,
            "commissionPaid": 0.0,
            "sources": set(),
            "outlets": set(),
            "lastVisitIdx": -1,
            "lastVisit": "",
            "phone": "",
            "doctors": set(),
            "treatments": [],      # [{date, item, code, qty, spend, outlet, doctor}]
            "firstDateISO": "",
            "lastDateISO": "",
        })
        if name_clean and name_clean.upper() != "NONE":
            c["names"].add(name_clean)
            if not c["name"] or len(name_clean) > len(c["name"]):
                c["name"] = name_clean
        if source:
            c["sources"].add(source)
        return mrn

    def add_invoice(self, mrn, name, amount, outlet, month_idx, month_label, source="invoice"):
        key = self.touch(mrn, name, source)
        if not key:
            return
        c = self._by_mrn[key]
        c["revenue"] += amount
        c["invoiceVisits"] += 1
        if outlet:
            c["outlets"].add(outlet)
        if month_idx > c["lastVisitIdx"]:
            c["lastVisitIdx"] = month_idx
            c["lastVisit"] = month_label

    def add_line_item(self, mrn, name, outlet="", source=""):
        key = self.touch(mrn, name, source)
        if not key:
            return
        c = self._by_mrn[key]
        c["lineItems"] += 1
        if outlet:
            c["outlets"].add(outlet)

    def set_phone(self, mrn, phone, name=""):
        if not phone:
            return
        key = self.touch(mrn, name)
        if key and not self._by_mrn[key]["phone"]:
            self._by_mrn[key]["phone"] = str(phone).strip()

    def add_treatment(self, mrn, name, item, code, qty, spend, date_obj, outlet="", doctor="", source="invoice"):
        """Record one dated treatment line for a patient (powers the date-range query)."""
        key = self.touch(mrn, name, source)
        if not key or not item:
            return
        c = self._by_mrn[key]
        iso = date_obj.strftime("%Y-%m-%d") if date_obj else ""
        c["treatments"].append({
            "date": iso,
            "item": str(item).strip(),
            "code": str(code).strip(),
            "qty": int(qty) if qty else 0,
            "spend": round(spend) if spend else 0,
            "outlet": outlet_label(outlet) if outlet else "",
            "doctor": str(doctor).strip() if doctor and str(doctor).upper() != "NONE" else "",
        })
        if outlet:
            c["outlets"].add(outlet_label(outlet))
        if doctor and str(doctor).upper() != "NONE":
            c["doctors"].add(str(doctor).strip())
        if iso:
            if not c["firstDateISO"] or iso < c["firstDateISO"]:
                c["firstDateISO"] = iso
            if not c["lastDateISO"] or iso > c["lastDateISO"]:
                c["lastDateISO"] = iso

    def add_commission(self, mrn, name, commission, source="commission"):
        key = self.touch(mrn, name, source)
        if not key:
            return
        c = self._by_mrn[key]
        c["commissionUsages"] += 1
        c["commissionPaid"] += commission

    def set_last_visit_label(self, mrn, month_label, month_rank):
        key = normalize_mrn(mrn)
        if key not in self._by_mrn:
            self.touch(key, source="historical")
        c = self._by_mrn.get(key)
        if not c:
            return
        rank = month_rank.get(month_label, -1)
        if rank > c["lastVisitIdx"]:
            c["lastVisitIdx"] = rank
            c["lastVisit"] = month_label

    def finalize(self):
        out = []
        ambiguous_names = []
        for nk, mrns in self._name_mrns.items():
            if len(mrns) > 1:
                ambiguous_names.append({
                    "nameKey": nk,
                    "mrns": sorted(mrns),
                    "count": len(mrns),
                })

        for mrn, c in self._by_mrn.items():
            names = sorted(c["names"])
            display = c["name"] or (names[0] if names else mrn)
            out.append({
                "mrn": mrn,
                "name": display,
                "aliases": names[:8],
                "revenue": round(c["revenue"]),
                "invoiceVisits": int(c["invoiceVisits"]),
                "lineItems": int(c["lineItems"]),
                "commissionUsages": int(c["commissionUsages"]),
                "commissionPaid": round(c["commissionPaid"]),
                "sources": sorted(c["sources"]),
                "outlets": sorted(c["outlets"]),
                "lastVisit": c["lastVisit"],
            })
        out.sort(key=lambda x: x["revenue"], reverse=True)
        return out, ambiguous_names

    def patient_last_map(self):
        return {mrn: c["lastVisit"] for mrn, c in self._by_mrn.items() if c["lastVisit"]}

    def patient_outlets(self):
        return {mrn: set(c["outlets"]) for mrn, c in self._by_mrn.items()}

    def visit_counts(self):
        return {mrn: c["invoiceVisits"] for mrn, c in self._by_mrn.items()}

    def patient_profiles(self):
        """Full per-patient profiles with dated treatments — powers patient-detail,
        remarketing, RFM, cohorts, and the date-range treatment query."""
        profiles = []
        for mrn, c in self._by_mrn.items():
            outs = {o for o in c["outlets"] if o in ("SpDVE", "SpGK")}
            if len(outs) >= 2:
                outlet = "Both"
            elif outs:
                outlet = next(iter(outs))
            else:
                outlet = ""
            txs = sorted(c["treatments"], key=lambda t: t["date"] or "9999")
            total_spend = round(c["revenue"]) if c["revenue"] else sum(t["spend"] for t in txs)
            visits = int(c["invoiceVisits"]) or len({t["date"] for t in txs if t["date"]})
            names = sorted(c["names"])
            profiles.append({
                "mrn": mrn,
                "name": c["name"] or (names[0] if names else mrn),
                "phone": c["phone"],
                "outlet": outlet,
                "totalSpend": total_spend,
                "visits": visits,
                "firstVisitDate": c["firstDateISO"],
                "lastVisitDate": c["lastDateISO"],
                "lastVisit": c["lastVisit"],
                "doctors": sorted(c["doctors"]),
                "treatments": txs,
            })
        profiles.sort(key=lambda p: p["totalSpend"], reverse=True)
        return profiles


def infer_data_year(invoice_months):
    """Derive the dataset's year from the invoice dates, instead of hardcoding it.
    Picks the most common year seen across the Collection-by-Invoice files."""
    counts = {}
    for _short, _label, _key, fp in invoice_months:
        try:
            for row in read_csv(fp):
                d = parse_invoice_date(row.get("Invoice Date", ""))
                if d:
                    counts[d.year] = counts.get(d.year, 0) + 1
        except Exception:
            pass
    if counts:
        return max(counts, key=counts.get)
    return datetime.now().year


def discover_files():
    """Return inventory of all raw CSVs and which months exist per type."""
    inventory = []
    months_by_type = defaultdict(list)
    for path in sorted(RAW_DIR.glob("*.csv")):
        inventory.append(path.name)
    for short, label, key in MONTH_ORDER:
        for ftype, pattern in FILE_TYPES.items():
            fp = RAW_DIR / pattern.format(m=short)
            if fp.is_file():
                months_by_type[ftype].append((short, label, key, fp))
    return inventory, months_by_type


def load_historical_patients(registry, data_year):
    """Optional: extend customer last-visit from historical XLSX.
    Ranks prior calendar years below the current dataset year (which the
    invoice loop ranks by month index)."""
    if not HISTORICAL_XLSX.is_file():
        return False
    try:
        import openpyxl
    except ImportError:
        return False

    # Years before the dataset year, oldest = lowest rank. The current data_year is
    # intentionally absent so historical rows never overwrite a precise month label.
    month_rank = {str(y): (y - 2019) for y in range(2019, data_year)}

    wb = openpyxl.load_workbook(HISTORICAL_XLSX, read_only=True, data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = ws.iter_rows(values_only=True)
        hdr = [str(c).strip() if c else "" for c in next(rows)]

        def ci(n, alt=None):
            try:
                return hdr.index(n)
            except ValueError:
                if alt:
                    try:
                        return hdr.index(alt)
                    except ValueError:
                        pass
            return None

        ii, mi, ni, dti = ci("Invoice Code"), ci("Nomor Rekam Medis", "MRN"), ci("Patient"), ci("Invoice Date")
        if mi is None:
            continue
        for row in rows:
            if not row:
                continue
            inv = str(row[ii]).strip() if ii is not None and row[ii] else ""
            if not inv or inv == "None":
                continue
            mrn = normalize_mrn(row[mi] if row[mi] else "")
            if not mrn:
                continue
            name = str(row[ni]).strip() if ni is not None and row[ni] else ""
            date = str(row[dti]).strip() if dti is not None and row[dti] else ""
            yr = None
            for p in date.replace(",", " ").split():
                if len(p) == 4 and p.isdigit() and p.startswith("20"):
                    yr = p
                    break
            if not yr and len(mrn) >= 5:
                yy = mrn[3:5]
                if yy.isdigit():
                    yr = "20" + yy
            if yr and yr in month_rank:
                registry.touch(mrn, name, "historical")
                registry.set_last_visit_label(mrn, yr, month_rank)
    wb.close()
    return True


def build():
    inventory, months_by_type = discover_files()
    invoice_months = months_by_type.get("invoice", [])
    if not invoice_months:
        raise FileNotFoundError(f"No Collection by Invoice CSV files in {RAW_DIR}")

    monthly = []
    items_acc = {}
    doctors_acc = {}
    payment_acc = {}
    dow_acc = {d: 0.0 for d in DAY_NAMES}
    credit_total = credit_invoices = 0
    credit_redemptions = credit_purchases = 0
    latest_date = None
    source_files = {"invoice": [], "collection": [], "collection_list": [], "sales": [], "sales_list": [], "commission": []}

    discount_total = discounted_invoices = gross_total = 0
    registry = CustomerRegistry()
    month_visitors = defaultdict(set)
    month_new = defaultdict(set)

    prev_revenue = None
    now = datetime.now()
    data_year = infer_data_year(invoice_months)

    # ── Collection by Invoice (primary) ───────────────────────────────────
    for month_idx, (short, label, key, fp) in enumerate(invoice_months):
        source_files["invoice"].append(fp.name)
        rows = read_csv(fp)
        spdve = spgk = other = 0.0
        inv_count = 0
        month_credit_v = 0.0
        month_label = f"{label} {data_year}"

        for row in rows:
            inv_code = str(row.get("Invoice Code", "")).strip()
            pm = row.get("Payment Method", "")
            ic = row.get("Item Code", "")
            outlet = str(row.get("Outlet", "")).strip()
            item_name = str(row.get("Item", "")).strip()
            doctor = str(row.get("Doctor", "")).strip()
            mrn, patient_name_row = extract_patient_fields(row)
            line_total = parse_amount(row.get("Total", 0))
            price = parse_amount(row.get("Price", 0))
            disc_val = parse_amount(row.get("Disc in Value", 0))
            disc_pct = parse_pct(row.get("Disc in %", 0))
            tx_date = parse_invoice_date(row.get("Invoice Date", ""))
            phone_row = extract_phone(row)
            if mrn and phone_row:
                registry.set_phone(mrn, phone_row, patient_name_row)

            if inv_code:
                amt = parse_amount(row.get("Total Paid", 0))
                inv_date = parse_invoice_date(row.get("Invoice Date", ""))
                gross_total += price if price else amt

                if is_credit(pm, ic):
                    credit_invoices += 1
                    month_credit_v += amt
                    if "CREDIT VOUCHER" in str(pm).upper():
                        credit_redemptions += 1
                    if str(ic).upper().startswith("CV-"):
                        credit_purchases += 1
                    continue
                if is_internal(pm):
                    continue

                inv_count += 1
                ol = outlet_label(outlet)
                if ol == "SpDVE":
                    spdve += amt
                elif ol == "SpGK":
                    spgk += amt
                else:
                    other += amt

                if disc_val > 0 or disc_pct > 0:
                    discounted_invoices += 1
                    discount_total += disc_val if disc_val else price * disc_pct / 100

                if doctor and doctor.upper() != "NONE":
                    d = doctors_acc.setdefault(doctor, {"name": doctor, "outlet": ol, "revenue": 0.0, "txns": 0})
                    d["revenue"] += amt
                    d["txns"] += 1
                    if ol in ("SpDVE", "SpGK"):
                        d["outlet"] = ol

                if mrn:
                    registry.add_invoice(mrn, patient_name_row, amt, outlet, month_idx, month_label)
                    month_visitors[key].add(mrn)
                    if mrn_new_month(mrn, key):
                        month_new[key].add(mrn)

                pay_key = normalize_payment(pm)
                payment_acc[pay_key] = payment_acc.get(pay_key, 0.0) + amt

                if inv_date:
                    if latest_date is None or inv_date > latest_date:
                        latest_date = inv_date
                    dow_acc[DAY_NAMES[inv_date.weekday()]] += amt

            if item_name and ic and not is_credit(pm, ic) and not is_internal(pm):
                ol = outlet_label(outlet)
                ik = (ic, item_name, ol)
                it = items_acc.setdefault(ik, {"item": item_name, "code": ic, "outlet": ol, "revenue": 0.0, "qty": 0, "txns": 0, "category": item_category(ic, item_name)})
                qty = int(parse_amount(row.get("Qty", 1)) or 1)
                it["revenue"] += line_total if line_total else 0
                it["qty"] += max(qty, 1)
                it["txns"] += 1
                if mrn:
                    registry.add_line_item(mrn, patient_name_row, outlet, "invoice")
                    registry.add_treatment(mrn, patient_name_row, item_name, ic,
                                           qty, line_total, tx_date, outlet, doctor, "invoice")

        credit_total += month_credit_v
        revenue = spdve + spgk + other
        aov = revenue / inv_count if inv_count else 0
        mom = None if prev_revenue is None else round((revenue - prev_revenue) / prev_revenue * 100, 1) if prev_revenue else 0
        prev_revenue = revenue

        display_label = label
        if key == invoice_months[-1][2] and latest_date and latest_date.month == now.month and latest_date.year == now.year:
            display_label = label + "*"

        monthly.append({
            "month": display_label,
            "key": key,
            "revenue": round(revenue),
            "txns": inv_count,
            "aov": round(aov),
            "mom": 0 if len(monthly) == 0 else mom,
            "spdve": round(spdve),
            "spgk": round(spgk),
        })

    if monthly:
        monthly[0]["mom"] = 0

    has_xlsx = load_historical_patients(registry, data_year)
    patient_last = registry.patient_last_map()
    patient_outlet = registry.patient_outlets()
    visit_counts = registry.visit_counts()

    # Patient lifecycle (dynamic active = last 2 invoice months)
    month_labels_year = [f"{m[1]} {data_year}" for m in invoice_months]
    active_set = set(month_labels_year[-2:]) if len(month_labels_year) >= 2 else set(month_labels_year)
    lapsing_set = set(month_labels_year[:-2]) if len(month_labels_year) > 2 else set()

    active_mrns = {m for m, v in patient_last.items() if v in active_set}
    lapsing_mrns = {m for m, v in patient_last.items() if v in lapsing_set}
    dormant_mrns = {m for m, v in patient_last.items() if v not in active_set and v not in lapsing_set}
    total_patients = len(patient_last)

    def outlet_count(mrns):
        out = defaultdict(int)
        for mrn in mrns:
            for o in patient_outlet.get(mrn, {"Unknown"}):
                out[o] += 1
        return dict(out)

    lapsing_breakdown = defaultdict(int)
    for mrn in lapsing_mrns:
        lapsing_breakdown[patient_last[mrn]] += 1

    new_returning = []
    for short, label, key, _ in invoice_months:
        visitors = month_visitors.get(key, set())
        new_ct = len(month_new.get(key, set()))
        ret_ct = max(0, len(visitors) - new_ct)
        new_returning.append({"month": label, "key": key, "new": new_ct, "returning": ret_ct})

    visit_buckets = {"1 visit": 0, "2 visits": 0, "3–5": 0, "6–10": 0, "11+": 0}
    for mrn, cnt in visit_counts.items():
        if cnt == 1:
            visit_buckets["1 visit"] += 1
        elif cnt == 2:
            visit_buckets["2 visits"] += 1
        elif cnt <= 5:
            visit_buckets["3–5"] += 1
        elif cnt <= 10:
            visit_buckets["6–10"] += 1
        else:
            visit_buckets["11+"] += 1
    visit_dist = [{"bucket": k, "count": v} for k, v in visit_buckets.items()]

    # ── Item Sold by Collection (categories + validation) ─────────────────────
    categories_acc = defaultdict(float)
    collection_monthly = []
    for short, label, key, fp in months_by_type.get("collection", []):
        source_files["collection"].append(fp.name)
        total = 0.0
        for row in read_csv(fp):
            cat = item_category(row.get("Item Code", ""), row.get("Item name", ""), row.get("Category", ""))
            val = parse_amount(row.get("Item Total", 0))
            categories_acc[cat] += val
            total += val
        collection_monthly.append({"month": label, "key": key, "total": round(total)})

    categories = [
        {"category": k, "revenue": round(v)}
        for k, v in sorted(categories_acc.items(), key=lambda x: -x[1])
    ]

    # ── Item Sold by Collection List (outlet-aware line items) ───────────────
    for short, label, key, fp in months_by_type.get("collection_list", []):
        source_files["collection_list"].append(fp.name)
        for row in read_csv(fp):
            ic = str(row.get("Item Code", "")).strip()
            item_name = str(row.get("Item", "")).strip()
            outlet = str(row.get("Outlet", "")).strip()
            line_total = parse_amount(row.get("Total", 0))
            mrn, pname = extract_patient_fields(row)
            if not item_name or not ic:
                continue
            pm = row.get("Payment Method", "")
            if is_credit(pm, ic) or is_internal(pm):
                continue
            ol = outlet_label(outlet)
            ik = (ic, item_name, ol)
            it = items_acc.setdefault(ik, {
                "item": item_name, "code": ic, "outlet": ol, "revenue": 0.0, "qty": 0, "txns": 0,
                "category": item_category(ic, item_name),
            })
            qty = int(parse_amount(row.get("Qty", 1)) or 1)
            it["revenue"] += line_total
            it["qty"] += max(qty, 1)
            it["txns"] += 1
            if mrn:
                registry.add_line_item(mrn, pname, outlet, "collection_list")

    # ── Item Sold by Sales (accrual comparison) ───────────────────────────────
    sales_monthly = []
    for short, label, key, fp in months_by_type.get("sales", []):
        source_files["sales"].append(fp.name)
        total = sum(parse_amount(r.get("Item Total", 0)) for r in read_csv(fp))
        sales_monthly.append({"month": label, "key": key, "total": round(total)})

    for short, label, key, fp in months_by_type.get("sales_list", []):
        source_files["sales_list"].append(fp.name)
        for row in read_csv(fp):
            mrn, pname = extract_patient_fields(row)
            ic = str(row.get("Item Code", "")).strip()
            if not mrn or not ic:
                continue
            pm = row.get("Payment Method", "")
            if is_credit(pm, ic) or is_internal(pm):
                continue
            outlet = str(row.get("Outlet", "")).strip()
            registry.add_line_item(mrn, pname, outlet, "sales_list")

    sales_vs_collection = []
    coll_map = {c["key"]: c["total"] for c in collection_monthly}
    for s in sales_monthly:
        coll = coll_map.get(s["key"], 0)
        sales_vs_collection.append({
            "month": s["month"],
            "key": s["key"],
            "sales": s["total"],
            "collection": coll,
            "invoice": next((m["revenue"] for m in monthly if m["key"] == s["key"]), 0),
        })

    # ── Commission Sales Usage Report ─────────────────────────────────────────
    commission_total = 0.0
    commission_usages = 0
    therapists_acc = defaultdict(lambda: {"name": "", "commission": 0.0, "usages": 0})
    for short, label, key, fp in months_by_type.get("commission", []):
        source_files["commission"].append(fp.name)
        for row in read_csv(fp):
            comm = parse_amount(row.get("Total Commission", 0))
            commission_total += comm
            commission_usages += 1
            mrn, pname = extract_patient_fields(row)
            if mrn:
                registry.add_commission(mrn, pname, comm)
            th = str(row.get("Therapist", "")).strip()
            if th and th.upper() != "NONE":
                t = therapists_acc[th]
                t["name"] = th
                t["commission"] += comm
                t["usages"] += 1

    therapists = sorted(therapists_acc.values(), key=lambda x: x["commission"], reverse=True)[:10]
    for t in therapists:
        t["commission"] = round(t["commission"])
        t["usages"] = int(t["usages"])

    # ── Finalize core aggregates ──────────────────────────────────────────────
    items = sorted(items_acc.values(), key=lambda x: x["revenue"], reverse=True)[:20]
    for it in items:
        it["revenue"] = round(it["revenue"])
        it["qty"] = int(it["qty"])
        it["txns"] = int(it["txns"])

    doctors = sorted(doctors_acc.values(), key=lambda x: x["revenue"], reverse=True)[:10]
    for d in doctors:
        d["revenue"] = round(d["revenue"])
        d["txns"] = int(d["txns"])

    customers_compiled, ambiguous_names = registry.finalize()
    patient_profiles = registry.patient_profiles()
    tx_total = sum(len(p["treatments"]) for p in patient_profiles)
    tx_missing_date = sum(1 for p in patient_profiles for t in p["treatments"] if not t.get("date"))
    patients = [
        {
            "rank": i + 1,
            "mrn": c["mrn"],
            "name": c["name"],
            "visits": c["invoiceVisits"],
            "revenue": c["revenue"],
            "lineItems": c["lineItems"],
            "commissionUsages": c["commissionUsages"],
            "sources": c["sources"],
        }
        for i, c in enumerate(customers_compiled[:15])
    ]

    pay_total = sum(payment_acc.values()) or 1
    payment = [
        {"method": k, "pct": round(v / pay_total * 100, 1)}
        for k, v in sorted(payment_acc.items(), key=lambda x: -x[1])[:8]
    ]

    dow = [{"day": d, "revenue": round(dow_acc[d])} for d in DAY_NAMES]

    grand_rev = sum(m["revenue"] for m in monthly)
    grand_spdve = sum(m["spdve"] for m in monthly)
    grand_spgk = sum(m["spgk"] for m in monthly)

    disc_pct = round(discounted_invoices / sum(m["txns"] for m in monthly) * 100, 1) if monthly else 0

    # Dynamic alerts from data
    alerts = []
    if len(monthly) >= 2:
        worst = min(monthly, key=lambda m: m["revenue"])
        best = max(monthly, key=lambda m: m["revenue"])
        if worst.get("mom") and worst["mom"] < -5:
            alerts.append({"type": "danger", "title": f"{worst['month']} Revenue Drop {worst['mom']}%",
                           "body": f"Lowest month at Rp {worst['revenue']:,.0f}."})
        if best.get("mom") and best["mom"] > 10:
            alerts.append({"type": "success", "title": f"{best['month']} Revenue Spike +{best['mom']}%",
                           "body": f"Best month at Rp {best['revenue']:,.0f}."})
    if payment:
        top_pay = payment[0]
        alerts.append({"type": "info", "title": f"{top_pay['method']} Dominates at {top_pay['pct']}%",
                       "body": "Top payment rail from invoice data."})
    alerts.append({"type": "purple",
                   "title": f"SpDVE {round(grand_spdve/grand_rev*100,1) if grand_rev else 0}% · SpGK {round(grand_spgk/grand_rev*100,1) if grand_rev else 0}%",
                   "body": f"All {len(inventory)} raw CSV files analysed."})

    first_month = monthly[0]["month"].replace("*", "")
    last_month = monthly[-1]["month"]

    return {
        "meta": {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "monthsLoaded": len(monthly),
            "dataYear": data_year,
            "treatmentsTotal": tx_total,
            "treatmentsMissingDate": tx_missing_date,
            "monthRange": f"{first_month} – {last_month}",
            "latestInvoiceDate": latest_date.strftime("%d %b %Y") if latest_date else None,
            "rawFilesCount": len(inventory),
            "rawFiles": inventory,
            "sourceFiles": source_files,
            "creditExcluded": round(credit_total),
            "creditInvoices": credit_invoices,
            "creditRedemptions": credit_redemptions,
            "creditPurchases": credit_purchases,
            "outletSpdvePct": round(grand_spdve / grand_rev * 100, 1) if grand_rev else 0,
            "outletSpgkPct": round(grand_spgk / grand_rev * 100, 1) if grand_rev else 0,
            "rawDataDir": str(RAW_DIR),
            "historicalPatientsLoaded": has_xlsx,
            "dataTypesLoaded": {k: len(v) for k, v in source_files.items()},
            "uniqueCustomers": len(customers_compiled),
            "customersCompiledFrom": "MRN (Nomor Rekam Medis) — same ID merged across invoice, collection list, sales list, commission",
            "ambiguousNameCount": len(ambiguous_names),
        },
        "monthly": monthly,
        "items": items,
        "doctors": doctors,
        "patients": patients,
        "customers": customers_compiled[:200],
        "patientProfiles": patient_profiles,
        "ambiguousNames": ambiguous_names[:20],
        "payment": payment,
        "dow": dow,
        "categories": categories,
        "newReturning": new_returning,
        "visitDist": visit_dist,
        "patientsLifecycle": {
            "total": total_patients,
            "active": len(active_mrns),
            "lapsing": len(lapsing_mrns),
            "dormant": len(dormant_mrns),
            "activePct": round(len(active_mrns) / total_patients * 100, 1) if total_patients else 0,
            "lapsingPct": round(len(lapsing_mrns) / total_patients * 100, 1) if total_patients else 0,
            "dormantPct": round(len(dormant_mrns) / total_patients * 100, 1) if total_patients else 0,
            "activeOutlets": outlet_count(active_mrns),
            "lapsingBreakdown": dict(lapsing_breakdown),
            "activeMonths": list(active_set),
            "hasHistorical": has_xlsx,
        },
        "discounts": {
            "totalDiscount": round(discount_total),
            "discountedInvoices": discounted_invoices,
            "discountedPct": disc_pct,
            "grossTotal": round(gross_total),
        },
        "salesVsCollection": sales_vs_collection,
        "commission": {
            "total": round(commission_total),
            "usages": commission_usages,
            "therapists": therapists,
        },
        "alerts": alerts,
    }


def main():
    try:
        payload = build()
        with open(OUT_PATH, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        m = payload["meta"]
        types = m.get("dataTypesLoaded", {})
        print(f"OK — {m['monthsLoaded']} months ({m['monthRange']}) · {m['rawFilesCount']} CSV files -> {OUT_PATH}")
        print(f"    Types: invoice={types.get('invoice',0)} collection={types.get('collection',0)} "
              f"list={types.get('collection_list',0)} sales={types.get('sales',0)} "
              f"commission={types.get('commission',0)}")
        if m.get("latestInvoiceDate"):
            print(f"    Latest invoice: {m['latestInvoiceDate']}")
        if m.get("historicalPatientsLoaded"):
            print("    Historical XLSX loaded for patient lifecycle")
        print(f"    Unique customers compiled (MRN): {m.get('uniqueCustomers', 0):,}")
        return 0
    except Exception as e:
        err = {"error": str(e), "trace": traceback.format_exc(), "generatedAt": datetime.now(timezone.utc).isoformat()}
        with open(OUT_PATH, "w", encoding="utf-8") as f:
            json.dump(err, f, indent=2)
        print(f"ERROR: {e}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
